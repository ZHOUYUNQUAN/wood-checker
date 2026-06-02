"""
木材检尺对比系统 - 路由逻辑
处理 Excel 上传、数据检索、材积计算等核心功能
"""
import os
import math
from datetime import datetime
from pathlib import Path

from flask import render_template, request, jsonify, current_app
from werkzeug.utils import secure_filename
import openpyxl

from models import get_db, _validate_table_name
from . import check_bp

# ========== 常量 ==========
ALLOWED_EXTENSIONS = {'xlsx', 'xlsm'}

# 需要跳过汇总行的关键词
SKIP_KEYWORDS = ('总计', '合计', 'Grand Total', '小计', 'Subtotal')

# 预设列标签候选（按顺序展示，长度 12）
PRESET_LABELS = [
    '顺序编号', '材种编码', '英文代码',
    '直径1', '直径2', '直径3', '直径4',
    '直径(CM)', '长度(M)', '材积(M³)', '客户', '是否转口'
]


def _allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _make_safe_upload_name(filename):
    """生成适合保存到上传目录的文件名，兼容中文原始文件名。"""
    ext = Path(filename).suffix.lower()
    stem = Path(filename).stem
    safe_stem = secure_filename(stem) or datetime.now().strftime('upload_%Y%m%d_%H%M%S')
    return f'{safe_stem}{ext}'


def _safe_upload_path(file_key):
    """返回上传文件绝对路径；非法 file_key 返回 None。"""
    if not file_key or os.path.basename(file_key) != file_key:
        return None
    if secure_filename(Path(file_key).stem) != Path(file_key).stem:
        return None
    if Path(file_key).suffix.lower().lstrip('.') not in ALLOWED_EXTENSIONS:
        return None
    return os.path.join(current_app.config['UPLOAD_FOLDER'], file_key)


def _safe_float(value, default=0.0):
    """安全转换为 float，非数值返回默认值"""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def _read_headers(ws):
    """读取 Excel 第一行作为表头列表"""
    row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = []
    for cell in row:
        if cell is not None:
            headers.append(str(cell).strip())
        else:
            headers.append('')
    return headers


def _parse_row(row_data, col_map):
    """
    解析 Excel 一行数据，返回字典或 None（需跳过的行）
    col_map: {field_name: column_index}，例如 {'no': 0, 'especie': 1, ...}
    """
    # 检查是否为汇总行
    for keyword in SKIP_KEYWORDS:
        for cell in row_data:
            if cell is not None and keyword in str(cell):
                return None

    # 检查是否全空行
    if all(cell is None for cell in row_data):
        return None

    no_idx = col_map.get('no')
    if no_idx is None:
        return None
    no_val = row_data[no_idx] if no_idx < len(row_data) else None
    no = str(no_val or '').strip()
    if not no:
        return None

    def _val(field):
        idx = col_map.get(field)
        if idx is not None and idx < len(row_data) and row_data[idx] is not None:
            return row_data[idx]
        return None

    def _str_val(field):
        return str(_val(field) or '').strip()

    return {
        'no': no,
        'especie': _str_val('especie'),
        'english_code': _str_val('english_code'),
        'diameter_1': _safe_float(_val('diameter_1')),
        'diameter_2': _safe_float(_val('diameter_2')),
        'diameter_3': _safe_float(_val('diameter_3')),
        'diameter_4': _safe_float(_val('diameter_4')),
        'diameter_avg': _safe_float(_val('diameter_avg')),
        'length_m': _safe_float(_val('length_m')),
        'volume_m3': _safe_float(_val('volume_m3')),
        'customer': _str_val('customer'),
        'is_transshipment': 1 if _str_val('is_transshipment') in (
            '是', 'Y', 'y', 'Yes', 'yes', '1', 'True', 'true'
        ) else 0,
    }


def _calc_external_standard(diameter, length):
    """外标公式：V = π × (D/100)² × L / 4"""
    return math.pi * (diameter / 100) ** 2 * length / 4


def _calc_national_standard(diameter, length):
    """
    国标公式（转自 Excel IF 嵌套），结果保留 4 位小数
    条件1: D <= 12 && L <= 10
    条件2: L <= 10 && D >= 14
    条件3: L >= 10.4
    条件4: L == 10.2 && D <= 12
    条件5: else（兜底：L == 10.2 && D > 12 及中间状态）
    """
    D = diameter
    L = length

    if D <= 12 and L <= 10:
        # 条件1
        return 0.7854 * L * (D + 0.45 * L + 0.2) ** 2 / 10000
    elif L <= 10 and D >= 14:
        # 条件2
        return 0.7854 * L * (D + 0.5 * L + 0.005 * L ** 2 + 0.000125 * L * (14 - L) ** 2 * (D - 10)) ** 2 / 10000
    elif L >= 10.4:
        # 条件3
        return 0.8 * L * (D + 0.5 * L) ** 2 / 10000
    elif L == 10.2 and D <= 12:
        # 条件4
        return (0.7854 * 10 * (D + 4.7) ** 2 / 10000) / 2 + (0.8 * 10.4 * (D + 5.2) ** 2 / 10000 / 2)
    else:
        # 条件5：兜底（L == 10.2 && D > 12，以及 D=13 且 L<=10 等中间态）
        return 0.8 * 10.4 * (D + 5.2) ** 2 / 20000 + 7.854 * (D + 5.5 + 0.02 * (D - 10)) ** 2 / 20000


# ========== 辅助：统一获取文件列表（registry + 旧 code_sheets） ==========

def _get_all_files():
    """返回合并后的文件列表：优先取 file_registry，再补 code_sheets 中未被迁移的旧文件"""
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT id, file_name, row_count, upload_time FROM file_registry ORDER BY upload_time DESC')
    registry_files = [dict(row) for row in cursor.fetchall()]
    registry_names = {f['file_name'] for f in registry_files}

    # 补上还在旧 code_sheets 中但未迁移到 registry 的文件
    if registry_names:
        placeholders = ','.join(['?' for _ in registry_names])
        cursor.execute(f'''
            SELECT file_name, COUNT(*) as row_count, MAX(upload_time) as upload_time
            FROM code_sheets
            WHERE file_name NOT IN ({placeholders})
            GROUP BY file_name
            ORDER BY MAX(upload_time) DESC
        ''', tuple(registry_names))
    else:
        cursor.execute('''
            SELECT file_name, COUNT(*) as row_count, MAX(upload_time) as upload_time
            FROM code_sheets
            GROUP BY file_name
            ORDER BY MAX(upload_time) DESC
        ''')
    legacy_files = [dict(row) for row in cursor.fetchall()]

    conn.close()
    return registry_files + legacy_files


def _resolve_table(cursor, file_name):
    """
    根据 file_name 查找对应的数据表名。
    先查 file_registry（独立表），找不到则假定在 code_sheets 中。
    返回 (table_name, in_registry) 或 (None, False)。
    """
    cursor.execute('SELECT table_name FROM file_registry WHERE file_name = ?', (file_name,))
    row = cursor.fetchone()
    if row:
        return row['table_name'], True
    return None, False


# ========== 路由 ==========

@check_bp.route('/')
def index():
    """主页：渲染 checker 页面，附带已上传文件列表"""
    files = _get_all_files()
    return render_template('check/checker.html', files=files, PREFIX='/check')


@check_bp.route('/upload', methods=['POST'])
def upload():
    """
    第一步：上传 Excel 文件，暂存并返回表头列名供用户选择编号列
    """
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': '未找到上传文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'ok': False, 'error': '文件名为空'}), 400

    if not _allowed_file(file.filename):
        return jsonify({'ok': False, 'error': '仅支持 .xlsx 和 .xlsm 文件'}), 400

    filename = _make_safe_upload_name(file.filename)
    upload_folder = current_app.config['UPLOAD_FOLDER']
    os.makedirs(upload_folder, exist_ok=True)

    # 防止不同原始文件名 secure_filename 后同名冲突：自动追加 _2, _3 ...
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM file_registry WHERE file_name = ?', (filename,))
    if cursor.fetchone():
        base, ext = os.path.splitext(filename)
        counter = 2
        while True:
            new_name = f'{base}_{counter}{ext}'
            cursor.execute('SELECT id FROM file_registry WHERE file_name = ?', (new_name,))
            if not cursor.fetchone() and not os.path.exists(os.path.join(upload_folder, new_name)):
                filename = new_name
                break
            counter += 1
    conn.close()

    file_path = _safe_upload_path(filename)
    if not file_path:
        return jsonify({'ok': False, 'error': '文件名无效'}), 400
    file.save(file_path)

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        headers = _read_headers(ws)

        sheets = []
        for name in wb.sheetnames:
            sheet = wb[name]
            row_count = max(sheet.max_row - 1, 0)
            sheets.append({'name': name, 'row_count': row_count})

        wb.close()
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Excel 解析失败: {str(e)}'}), 400

    if not any(h for h in headers):
        return jsonify({'ok': False, 'error': '未检测到表头行'}), 400

    return jsonify({
        'ok': True,
        'stage': 'preview',
        'file_key': filename,
        'sheets': sheets,
        'headers': headers,
        'message': f'已读取 {len(headers)} 列表头，请选择编号列',
    })


@check_bp.route('/upload/preview_sheet', methods=['POST'])
def preview_sheet():
    """获取指定 sheet 的表头数据，用于前端切换 sheet 时预览"""
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'error': '请求数据为空'}), 400

    file_key = data.get('file_key', '').strip()
    sheet_name = data.get('sheet_name', '').strip()

    if not file_key:
        return jsonify({'ok': False, 'error': '缺少 file_key'}), 400
    if not sheet_name:
        return jsonify({'ok': False, 'error': '缺少 sheet_name'}), 400

    upload_folder = current_app.config['UPLOAD_FOLDER']
    file_path = _safe_upload_path(file_key)
    if not file_path or not os.path.exists(file_path):
        return jsonify({'ok': False, 'error': '文件已失效，请重新上传'}), 400

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return jsonify({'ok': False, 'error': f'Sheet "{sheet_name}" 不存在'}), 400
        ws = wb[sheet_name]
        headers = _read_headers(ws)
        wb.close()
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Excel 解析失败: {str(e)}'}), 400

    return jsonify({
        'ok': True,
        'sheet_name': sheet_name,
        'headers': headers,
    })


@check_bp.route('/upload/confirm', methods=['POST'])
def upload_confirm():
    """
    第二步：根据用户指定的列映射执行完整解析并入库。
    每个上传文件创建独立数据表，通过 file_registry 记录元信息。
    """
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'error': '请求数据为空'}), 400

    file_key = data.get('file_key', '').strip()
    col_mapping = data.get('col_mapping')

    if not file_key:
        return jsonify({'ok': False, 'error': '缺少 file_key'}), 400
    if not col_mapping or col_mapping.get('no') is None:
        return jsonify({'ok': False, 'error': '请至少指定编号列'}), 400

    # 确保 value 都是整数
    try:
        col_mapping = {k: int(v) for k, v in col_mapping.items() if v is not None}
    except (ValueError, TypeError):
        return jsonify({'ok': False, 'error': '列索引格式错误'}), 400

    sheet_name = data.get('sheet_name', '').strip() or None

    upload_folder = current_app.config['UPLOAD_FOLDER']
    file_path = _safe_upload_path(file_key)
    if not file_path or not os.path.exists(file_path):
        return jsonify({'ok': False, 'error': '文件已失效，请重新上传'}), 400

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                wb.close()
                return jsonify({'ok': False, 'error': f'Sheet "{sheet_name}" 不存在'}), 400
            ws = wb[sheet_name]
        else:
            ws = wb.active
        headers = _read_headers(ws)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Excel 解析失败: {str(e)}'}), 400

    records = []
    for row in rows:
        parsed = _parse_row(row, col_mapping)
        if parsed is not None:
            records.append(parsed)

    if not records:
        return jsonify({'ok': False, 'error': '未找到有效数据行，请确认选择的编号列是否正确'}), 400

    # ---- 生成唯一表名 ----
    timestamp = datetime.now().strftime('%Y%m%d')
    rand_suffix = os.urandom(4).hex()
    table_name = f'sheets_{timestamp}_{rand_suffix}'

    conn = get_db()
    cursor = conn.cursor()

    # 如果 file_key 已在 registry 中（重新上传同名文件），先清理旧表
    cursor.execute('SELECT table_name FROM file_registry WHERE id = ?', (file_key,))
    old = cursor.fetchone()
    if old:
        _validate_table_name(old['table_name'])
        cursor.execute(f'DROP TABLE IF EXISTS "{old["table_name"]}"')

    _validate_table_name(table_name)
    # 创建独立数据表
    cursor.execute(f'''
        CREATE TABLE IF NOT EXISTS "{table_name}" (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            no TEXT,
            especie TEXT,
            english_code TEXT,
            diameter_1 REAL,
            diameter_2 REAL,
            diameter_3 REAL,
            diameter_4 REAL,
            diameter_avg REAL,
            length_m REAL,
            volume_m3 REAL,
            customer TEXT,
            is_transshipment INTEGER DEFAULT 0
        )
    ''')

    cursor.execute(f'''
        CREATE INDEX IF NOT EXISTS idx_{table_name}_no ON "{table_name}"(no)
    ''')

    upload_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # 批量插入数据
    for rec in records:
        cursor.execute(f'''
            INSERT INTO "{table_name}"
                (no, especie, english_code,
                 diameter_1, diameter_2, diameter_3, diameter_4,
                 diameter_avg, length_m, volume_m3, customer, is_transshipment)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            rec['no'], rec['especie'], rec['english_code'],
            rec['diameter_1'], rec['diameter_2'], rec['diameter_3'], rec['diameter_4'],
            rec['diameter_avg'], rec['length_m'], rec['volume_m3'],
            rec['customer'], rec['is_transshipment'],
        ))

    # 写入 registry
    cursor.execute('''
        INSERT OR REPLACE INTO file_registry (id, file_name, table_name, row_count, upload_time)
        VALUES (?, ?, ?, ?, ?)
    ''', (file_key, file_key, table_name, len(records), upload_time))

    # 同时清理旧 code_sheets 中可能存在的同名数据（向后兼容过渡期）
    cursor.execute('DELETE FROM code_sheets WHERE file_name = ?', (file_key,))

    conn.commit()
    conn.close()

    # 生成列映射摘要（字段名 → 表头文本）
    col_summary = {}
    for field, idx in col_mapping.items():
        if idx is not None and idx < len(headers):
            col_summary[field] = headers[idx]

    return jsonify({
        'ok': True,
        'file_name': file_key,
        'row_count': len(records),
        'col_map': col_summary,
        'message': f'成功导入 {len(records)} 条记录',
    })


@check_bp.route('/search')
def search():
    """模糊搜索 no 字段，返回匹配记录。file_name 指向独立表或旧表。"""
    q = request.args.get('q', '').strip()
    file_name = request.args.get('file_name', '').strip()

    if not q:
        return jsonify({'ok': False, 'error': '搜索关键词不能为空'}), 400

    conn = get_db()
    cursor = conn.cursor()

    if file_name:
        table_name, in_registry = _resolve_table(cursor, file_name)
        if in_registry and table_name:
            _validate_table_name(table_name)
            cursor.execute(f'SELECT * FROM "{table_name}" WHERE no LIKE ? ORDER BY no LIMIT 200', (f'%{q}%',))
        else:
            cursor.execute('''
                SELECT * FROM code_sheets
                WHERE no LIKE ? AND file_name = ?
                ORDER BY no
                LIMIT 200
            ''', (f'%{q}%', file_name))
    else:
        cursor.execute('''
            SELECT * FROM code_sheets
            WHERE no LIKE ?
            ORDER BY no
            LIMIT 200
        ''', (f'%{q}%',))

    records = [dict(row) for row in cursor.fetchall()]
    conn.close()

    return jsonify({'ok': True, 'records': records, 'count': len(records)})


@check_bp.route('/calc', methods=['POST'])
def calc():
    """计算新材积（国标或外标）。file_name 用于定位独立数据表。"""
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'error': '请求数据为空'}), 400

    record_id = data.get('record_id')
    file_name = data.get('file_name', '').strip()
    standard = data.get('standard', '').strip()
    length = data.get('length')
    diameter = data.get('diameter')

    if not record_id:
        return jsonify({'ok': False, 'error': '缺少 record_id'}), 400
    if standard not in ('national', 'external'):
        return jsonify({'ok': False, 'error': '标准参数无效，可选 national 或 external'}), 400

    conn = get_db()
    cursor = conn.cursor()

    if file_name:
        table_name, in_registry = _resolve_table(cursor, file_name)
        if in_registry and table_name:
            _validate_table_name(table_name)
            cursor.execute(f'SELECT * FROM "{table_name}" WHERE id = ?', (record_id,))
        else:
            cursor.execute('SELECT * FROM code_sheets WHERE id = ?', (record_id,))
    else:
        cursor.execute('SELECT * FROM code_sheets WHERE id = ?', (record_id,))

    row = cursor.fetchone()
    conn.close()

    if not row:
        return jsonify({'ok': False, 'error': '记录不存在'}), 404

    original_volume = row['volume_m3'] or 0.0

    # 使用传入的直径/长度，若未传则用数据库中的值
    if length is not None:
        try:
            length = float(length)
        except (ValueError, TypeError):
            length = row['length_m'] or 0.0
    else:
        length = row['length_m'] or 0.0

    if diameter is not None:
        try:
            diameter = float(diameter)
        except (ValueError, TypeError):
            diameter = row['diameter_avg'] or 0.0
    else:
        diameter = row['diameter_avg'] or 0.0

    # 根据标准选择公式计算
    if standard == 'external':
        new_volume = _calc_external_standard(diameter, length)
    else:
        new_volume = _calc_national_standard(diameter, length)

    new_volume = round(new_volume, 4)
    diff = round(new_volume - original_volume, 4)
    if original_volume != 0:
        rate = round((new_volume - original_volume) / original_volume * 100, 2)
    else:
        rate = 0.0

    return jsonify({
        'ok': True,
        'original_volume': original_volume,
        'new_volume': new_volume,
        'diff': diff,
        'rate': rate,
        'standard': standard,
        'diameter_used': diameter,
        'length_used': length,
    })


@check_bp.route('/files')
def get_files():
    """返回已上传文件列表（JSON），合并 registry 和旧 code_sheets"""
    files = _get_all_files()
    return jsonify({'ok': True, 'files': files})


@check_bp.route('/files/delete', methods=['POST'])
def delete_file():
    """删除文件记录及其所有数据（独立表则 DROP TABLE，否则删 code_sheets 行）"""
    data = request.get_json()
    if not data or not data.get('file_name'):
        return jsonify({'ok': False, 'error': '缺少 file_name'}), 400

    file_name = data['file_name']

    conn = get_db()
    cursor = conn.cursor()

    # 尝试 registry 路径
    cursor.execute('SELECT table_name FROM file_registry WHERE file_name = ?', (file_name,))
    reg_row = cursor.fetchone()

    if reg_row:
        table_name = reg_row['table_name']
        _validate_table_name(table_name)
        cursor.execute(f'SELECT COUNT(*) as cnt FROM "{table_name}"')
        count = cursor.fetchone()['cnt']
        cursor.execute(f'DROP TABLE IF EXISTS "{table_name}"')
        cursor.execute('DELETE FROM file_registry WHERE file_name = ?', (file_name,))
    else:
        # 旧表路径
        cursor.execute('SELECT COUNT(*) as cnt FROM code_sheets WHERE file_name = ?', (file_name,))
        count = cursor.fetchone()['cnt']
        cursor.execute('DELETE FROM code_sheets WHERE file_name = ?', (file_name,))

    conn.commit()
    conn.close()

    # 尝试删除上传目录中的文件（允许失败）
    upload_folder = current_app.config['UPLOAD_FOLDER']
    file_path = os.path.join(upload_folder, file_name)
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
    except OSError:
        pass

    return jsonify({'ok': True, 'deleted_count': count, 'message': f'已删除文件 {file_name} 及其 {count} 条数据'})


@check_bp.route('/files/rename', methods=['POST'])
def rename_file():
    """重命名 file_registry 中的文件（仅改显示名，不改底层表名和 id）"""
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'error': '请求数据为空'}), 400

    old_name = data.get('old_name', '').strip()
    new_name = data.get('new_name', '').strip()

    if not old_name:
        return jsonify({'ok': False, 'error': '缺少旧文件名'}), 400
    if not new_name:
        return jsonify({'ok': False, 'error': '新文件名不能为空'}), 400

    conn = get_db()
    cursor = conn.cursor()

    # 检查旧名是否存在
    cursor.execute('SELECT id FROM file_registry WHERE file_name = ?', (old_name,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'ok': False, 'error': f'文件 "{old_name}" 不存在，仅支持重命名新格式文件'}), 404

    # 检查新名是否冲突
    cursor.execute('SELECT id FROM file_registry WHERE file_name = ?', (new_name,))
    conflict = cursor.fetchone()
    if conflict and conflict['id'] != row['id']:
        conn.close()
        return jsonify({'ok': False, 'error': f'文件名 "{new_name}" 已存在'}), 409

    cursor.execute('UPDATE file_registry SET file_name = ? WHERE id = ?', (new_name, row['id']))
    conn.commit()
    conn.close()

    return jsonify({'ok': True, 'old_name': old_name, 'new_name': new_name, 'message': f'已重命名为 "{new_name}"'})
