"""
木材检尺对比系统 - 上传相关路由
处理 Excel 上传、文件管理（删除/重命名）等
"""
import os
from datetime import datetime

from flask import request, jsonify, current_app
import openpyxl

from models import get_db, _validate_table_name
from . import check_bp
from .utils import (
    allowed_file,
    make_safe_upload_name,
    safe_upload_path,
    read_headers,
    parse_row,
    get_all_files,
)


@check_bp.route('/upload', methods=['POST'])
def upload():
    """
    第一步：上传 Excel 文件，暂存并返回表头列名供用户选择编号列
    """
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': '未找到上传文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'ok': False, 'error': '文件名为空'}), 400

    if not allowed_file(file.filename):
        return jsonify({'ok': False, 'error': '仅支持 .xlsx 和 .xlsm 文件'}), 400

    filename = make_safe_upload_name(file.filename)
    upload_folder = current_app.config['UPLOAD_FOLDER']
    os.makedirs(upload_folder, exist_ok=True)

    # 防止不同原始文件名 secure_filename 后同名冲突：自动追加 _2, _3 ...
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM file_registry WHERE file_name = ?', (filename,))
    if cursor.fetchone():
        base, ext = os.path.splitext(filename)
        counter = 2
        while True:
            new_name = f'{base}_{counter}{ext}'
            cursor.execute('SELECT id FROM file_registry WHERE file_name = ?', (new_name,))
            if not cursor.fetchone() and not os.path.exists(os.path.join(upload_folder, new_name)):
                filename = new_name
                break
            counter += 1
    conn.close()

    file_path = safe_upload_path(filename)
    if not file_path:
        return jsonify({'ok': False, 'error': '文件名无效'}), 400
    file.save(file_path)

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        headers = read_headers(ws)

        sheets = []
        for name in wb.sheetnames:
            sheet = wb[name]
            row_count = max(sheet.max_row - 1, 0)
            sheets.append({'name': name, 'row_count': row_count})

        wb.close()
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Excel 解析失败: {str(e)}'}), 400

    if not any(h for h in headers):
        return jsonify({'ok': False, 'error': '未检测到表头行'}), 400

    return jsonify({
        'ok': True,
        'stage': 'preview',
        'file_key': filename,
        'sheets': sheets,
        'headers': headers,
        'message': f'已读取 {len(headers)} 列表头，请选择编号列',
    })


@check_bp.route('/upload/preview_sheet', methods=['POST'])
def preview_sheet():
    """获取指定 sheet 的表头数据，用于前端切换 sheet 时预览"""
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'error': '请求数据为空'}), 400

    file_key = data.get('file_key', '').strip()
    sheet_name = data.get('sheet_name', '').strip()

    if not file_key:
        return jsonify({'ok': False, 'error': '缺少 file_key'}), 400
    if not sheet_name:
        return jsonify({'ok': False, 'error': '缺少 sheet_name'}), 400

    upload_folder = current_app.config['UPLOAD_FOLDER']
    file_path = safe_upload_path(file_key)
    if not file_path or not os.path.exists(file_path):
        return jsonify({'ok': False, 'error': '文件已失效，请重新上传'}), 400

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return jsonify({'ok': False, 'error': f'Sheet "{sheet_name}" 不存在'}), 400
        ws = wb[sheet_name]
        headers = read_headers(ws)
        wb.close()
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Excel 解析失败: {str(e)}'}), 400

    return jsonify({
        'ok': True,
        'sheet_name': sheet_name,
        'headers': headers,
    })


@check_bp.route('/upload/confirm', methods=['POST'])
def upload_confirm():
    """
    第二步：根据用户指定的列映射执行完整解析并入库。
    每个上传文件创建独立数据表，通过 file_registry 记录元信息。
    """
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'error': '请求数据为空'}), 400

    file_key = data.get('file_key', '').strip()
    col_mapping = data.get('col_mapping')

    if not file_key:
        return jsonify({'ok': False, 'error': '缺少 file_key'}), 400
    if not col_mapping or col_mapping.get('no') is None:
        return jsonify({'ok': False, 'error': '请至少指定编号列'}), 400

    # 确保 value 都是整数
    try:
        col_mapping = {k: int(v) for k, v in col_mapping.items() if v is not None}
    except (ValueError, TypeError):
        return jsonify({'ok': False, 'error': '列索引格式错误'}), 400

    sheet_name = data.get('sheet_name', '').strip() or None

    upload_folder = current_app.config['UPLOAD_FOLDER']
    file_path = safe_upload_path(file_key)
    if not file_path or not os.path.exists(file_path):
        return jsonify({'ok': False, 'error': '文件已失效，请重新上传'}), 400

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                wb.close()
                return jsonify({'ok': False, 'error': f'Sheet "{sheet_name}" 不存在'}), 400
            ws = wb[sheet_name]
        else:
            ws = wb.active
        headers = read_headers(ws)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Excel 解析失败: {str(e)}'}), 400

    records = []
    for row in rows:
        parsed = parse_row(row, col_mapping)
        if parsed is not None:
            records.append(parsed)

    if not records:
        return jsonify({'ok': False, 'error': '未找到有效数据行，请确认选择的编号列是否正确'}), 400

    # ---- 生成唯一表名 ----
    timestamp = datetime.now().strftime('%Y%m%d')
    rand_suffix = os.urandom(4).hex()
    table_name = f'sheets_{timestamp}_{rand_suffix}'

    conn = get_db()
    cursor = conn.cursor()

    # 如果 file_key 已在 registry 中（重新上传同名文件），先清理旧表
    cursor.execute('SELECT table_name FROM file_registry WHERE id = ?', (file_key,))
    old = cursor.fetchone()
    if old:
        _validate_table_name(old['table_name'])
        cursor.execute(f'DROP TABLE IF EXISTS "{old["table_name"]}"')

    _validate_table_name(table_name)
    # 创建独立数据表
    cursor.execute(f'''
        CREATE TABLE IF NOT EXISTS "{table_name}" (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            no TEXT,
            especie TEXT,
            english_code TEXT,
            diameter_1 REAL,
            diameter_2 REAL,
            diameter_3 REAL,
            diameter_4 REAL,
            diameter_avg REAL,
            length_m REAL,
            volume_m3 REAL,
            customer TEXT,
            is_transshipment INTEGER DEFAULT 0
        )
    ''')

    cursor.execute(f'''
        CREATE INDEX IF NOT EXISTS idx_{table_name}_no ON "{table_name}"(no)
    ''')

    upload_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # 批量插入数据
    for rec in records:
        cursor.execute(f'''
            INSERT INTO "{table_name}"
                (no, especie, english_code,
                 diameter_1, diameter_2, diameter_3, diameter_4,
                 diameter_avg, length_m, volume_m3, customer, is_transshipment)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            rec['no'], rec['especie'], rec['english_code'],
            rec['diameter_1'], rec['diameter_2'], rec['diameter_3'], rec['diameter_4'],
            rec['diameter_avg'], rec['length_m'], rec['volume_m3'],
            rec['customer'], rec['is_transshipment'],
        ))

    # 写入 registry
    cursor.execute('''
        INSERT OR REPLACE INTO file_registry (id, file_name, table_name, row_count, upload_time)
        VALUES (?, ?, ?, ?, ?)
    ''', (file_key, file_key, table_name, len(records), upload_time))

    # 同时清理旧 code_sheets 中可能存在的同名数据（向后兼容过渡期）
    cursor.execute('DELETE FROM code_sheets WHERE file_name = ?', (file_key,))

    conn.commit()
    conn.close()

    # 生成列映射摘要（字段名 → 表头文本）
    col_summary = {}
    for field, idx in col_mapping.items():
        if idx is not None and idx < len(headers):
            col_summary[field] = headers[idx]

    return jsonify({
        'ok': True,
        'file_name': file_key,
        'row_count': len(records),
        'col_map': col_summary,
        'message': f'成功导入 {len(records)} 条记录',
    })


@check_bp.route('/files')
def list_files():
    """返回已上传文件列表（JSON），合并 registry 和旧 code_sheets"""
    files = get_all_files()
    return jsonify({'ok': True, 'files': files})


@check_bp.route('/files/delete', methods=['POST'])
def delete_file():
    """删除文件记录及其所有数据（独立表则 DROP TABLE，否则删 code_sheets 行）"""
    data = request.get_json()
    if not data or not data.get('file_name'):
        return jsonify({'ok': False, 'error': '缺少 file_name'}), 400

    file_name = data['file_name']

    conn = get_db()
    cursor = conn.cursor()

    # 尝试 registry 路径
    cursor.execute('SELECT table_name FROM file_registry WHERE file_name = ?', (file_name,))
    reg_row = cursor.fetchone()

    if reg_row:
        table_name = reg_row['table_name']
        _validate_table_name(table_name)
        cursor.execute(f'SELECT COUNT(*) as cnt FROM "{table_name}"')
        count = cursor.fetchone()['cnt']
        cursor.execute(f'DROP TABLE IF EXISTS "{table_name}"')
        cursor.execute('DELETE FROM file_registry WHERE file_name = ?', (file_name,))
    else:
        # 旧表路径
        cursor.execute('SELECT COUNT(*) as cnt FROM code_sheets WHERE file_name = ?', (file_name,))
        count = cursor.fetchone()['cnt']
        cursor.execute('DELETE FROM code_sheets WHERE file_name = ?', (file_name,))

    conn.commit()
    conn.close()

    # 尝试删除上传目录中的文件（允许失败）
    upload_folder = current_app.config['UPLOAD_FOLDER']
    file_path = os.path.join(upload_folder, file_name)
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
    except OSError:
        pass

    return jsonify({'ok': True, 'deleted_count': count, 'message': f'已删除文件 {file_name} 及其 {count} 条数据'})


@check_bp.route('/files/rename', methods=['POST'])
def rename_file():
    """重命名 file_registry 中的文件（仅改显示名，不改底层表名和 id）"""
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'error': '请求数据为空'}), 400

    old_name = data.get('old_name', '').strip()
    new_name = data.get('new_name', '').strip()

    if not old_name:
        return jsonify({'ok': False, 'error': '缺少旧文件名'}), 400
    if not new_name:
        return jsonify({'ok': False, 'error': '新文件名不能为空'}), 400

    conn = get_db()
    cursor = conn.cursor()

    # 检查旧名是否存在
    cursor.execute('SELECT id FROM file_registry WHERE file_name = ?', (old_name,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'ok': False, 'error': f'文件 "{old_name}" 不存在，仅支持重命名新格式文件'}), 404

    # 检查新名是否冲突
    cursor.execute('SELECT id FROM file_registry WHERE file_name = ?', (new_name,))
    conflict = cursor.fetchone()
    if conflict and conflict['id'] != row['id']:
        conn.close()
        return jsonify({'ok': False, 'error': f'文件名 "{new_name}" 已存在'}), 409

    cursor.execute('UPDATE file_registry SET file_name = ? WHERE id = ?', (new_name, row['id']))
    conn.commit()
    conn.close()

    return jsonify({'ok': True, 'old_name': old_name, 'new_name': new_name, 'message': f'已重命名为 "{new_name}"'})
